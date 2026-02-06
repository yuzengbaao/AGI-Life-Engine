"""Data Visualization Module.

Provides visualization capabilities for data analysis and reporting.
Supports:
- Statistical charts (histograms, box plots, violin plots)
- Comparison charts (bar, line, scatter)
- Distribution visualizations
- Time series plots
- Heatmaps and correlation matrices
- Multi-panel figures

Dependencies:
    - matplotlib: Core plotting library
    - seaborn: Statistical visualization
"""

import logging
from typing import Any, Dict, List, Optional, Tuple, Union

import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns

# Configure module logger
logger = logging.getLogger(__name__)

# Set style
sns.set_style("whitegrid")
plt.rcParams['figure.figsize'] = (10, 6)
plt.rcParams['font.size'] = 10


# ============================================================================
# Statistical Visualizations
# ============================================================================

class DataVisualizer:
    """Data visualization class for generating plots and charts.

    Example:
        >>> viz = DataVisualizer()
        >>> fig = viz.plot_histogram(df, 'age')
        >>> fig.savefig('histogram.png')
    """

    def __init__(self, style: str = "whitegrid"):
        """Initialize DataVisualizer.

        Args:
            style: Seaborn style (whitegrid, darkgrid, white, dark, ticks)
        """
        sns.set_style(style)
        self.style = style

    def plot_histogram(
        self,
        df: pd.DataFrame,
        column: str,
        bins: int = 30,
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (10, 6),
        **kwargs
    ) -> plt.Figure:
        """Plot histogram for a column.

        Args:
            df: Input DataFrame
            column: Column to plot
            bins: Number of bins
            title: Plot title
            figsize: Figure size (width, height)
            **kwargs: Additional matplotlib arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_histogram(df, 'age', bins=20)
        """
        logger.info(f"[VISUALIZE] Creating histogram for {column}")

        fig, ax = plt.subplots(figsize=figsize)

        df[column].plot(kind='hist', bins=bins, ax=ax, **kwargs)

        ax.set_xlabel(column)
        ax.set_ylabel('Frequency')
        ax.set_title(title or f'Distribution of {column}')

        plt.tight_layout()
        return fig

    def plot_boxplot(
        self,
        df: pd.DataFrame,
        column: str,
        by: Optional[str] = None,
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (10, 6),
        **kwargs
    ) -> plt.Figure:
        """Plot box plot for a column.

        Args:
            df: Input DataFrame
            column: Column to plot
            by: Grouping column (optional)
            title: Plot title
            figsize: Figure size
            **kwargs: Additional seaborn arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_boxplot(df, 'score', by='category')
        """
        logger.info(f"[VISUALIZE] Creating box plot for {column}")

        fig, ax = plt.subplots(figsize=figsize)

        if by:
            sns.boxplot(data=df, x=by, y=column, ax=ax, **kwargs)
        else:
            sns.boxplot(data=df, y=column, ax=ax, **kwargs)

        ax.set_title(title or f'Box Plot of {column}')
        plt.tight_layout()
        return fig

    def plot_violin(
        self,
        df: pd.DataFrame,
        x: str,
        y: str,
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (10, 6),
        **kwargs
    ) -> plt.Figure:
        """Plot violin plot.

        Args:
            df: Input DataFrame
            x: X-axis column
            y: Y-axis column
            title: Plot title
            figsize: Figure size
            **kwargs: Additional seaborn arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_violin(df, 'category', 'value')
        """
        logger.info(f"[VISUALIZE] Creating violin plot of {y} by {x}")

        fig, ax = plt.subplots(figsize=figsize)

        sns.violinplot(data=df, x=x, y=y, ax=ax, **kwargs)

        ax.set_title(title or f'Violin Plot of {y} by {x}')
        plt.tight_layout()
        return fig

    def plot_correlation_heatmap(
        self,
        df: pd.DataFrame,
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (12, 8),
        **kwargs
    ) -> plt.Figure:
        """Plot correlation matrix heatmap.

        Args:
            df: Input DataFrame
            title: Plot title
            figsize: Figure size
            **kwargs: Additional seaborn arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_correlation_heatmap(df)
        """
        logger.info("[VISUALIZE] Creating correlation heatmap")

        # Calculate correlation for numeric columns only
        numeric_df = df.select_dtypes(include=['number'])
        corr = numeric_df.corr()

        fig, ax = plt.subplots(figsize=figsize)

        sns.heatmap(
            corr,
            annot=True,
            fmt='.2f',
            cmap='coolwarm',
            center=0,
            ax=ax,
            **kwargs
        )

        ax.set_title(title or 'Correlation Matrix')
        plt.tight_layout()
        return fig

    def plot_scatter(
        self,
        df: pd.DataFrame,
        x: str,
        y: str,
        hue: Optional[str] = None,
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (10, 6),
        **kwargs
    ) -> plt.Figure:
        """Plot scatter plot.

        Args:
            df: Input DataFrame
            x: X-axis column
            y: Y-axis column
            hue: Grouping column (color)
            title: Plot title
            figsize: Figure size
            **kwargs: Additional seaborn arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_scatter(df, 'age', 'income', hue='education')
        """
        logger.info(f"[VISUALIZE] Creating scatter plot: {x} vs {y}")

        fig, ax = plt.subplots(figsize=figsize)

        sns.scatterplot(data=df, x=x, y=y, hue=hue, ax=ax, **kwargs)

        ax.set_title(title or f'Scatter Plot: {x} vs {y}')
        plt.tight_layout()
        return fig

    def plot_bar(
        self,
        df: pd.DataFrame,
        x: str,
        y: str,
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (10, 6),
        **kwargs
    ) -> plt.Figure:
        """Plot bar chart.

        Args:
            df: Input DataFrame
            x: X-axis column
            y: Y-axis column
            title: Plot title
            figsize: Figure size
            **kwargs: Additional seaborn arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_bar(df, 'category', 'sales')
        """
        logger.info(f"[VISUALIZE] Creating bar chart: {y} by {x}")

        fig, ax = plt.subplots(figsize=figsize)

        sns.barplot(data=df, x=x, y=y, ax=ax, **kwargs)

        ax.set_title(title or f'Bar Chart: {y} by {x}')
        plt.xticks(rotation=45)
        plt.tight_layout()
        return fig

    def plot_line(
        self,
        df: pd.DataFrame,
        x: str,
        y: str,
        hue: Optional[str] = None,
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (10, 6),
        **kwargs
    ) -> plt.Figure:
        """Plot line chart.

        Args:
            df: Input DataFrame
            x: X-axis column (often date/time)
            y: Y-axis column
            hue: Grouping column
            title: Plot title
            figsize: Figure size
            **kwargs: Additional seaborn arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_line(df, 'date', 'sales', hue='product')
        """
        logger.info(f"[VISUALIZE] Creating line chart: {y} over {x}")

        fig, ax = plt.subplots(figsize=figsize)

        sns.lineplot(data=df, x=x, y=y, hue=hue, ax=ax, **kwargs)

        ax.set_title(title or f'Line Chart: {y} over {x}')
        plt.xticks(rotation=45)
        plt.tight_layout()
        return fig

    def plot_time_series(
        self,
        df: pd.DataFrame,
        date_col: str,
        value_cols: Union[str, List[str]],
        title: Optional[str] = None,
        figsize: Tuple[int, int] = (12, 6),
        **kwargs
    ) -> plt.Figure:
        """Plot time series data.

        Args:
            df: Input DataFrame
            date_col: Date column name
            value_cols: Column(s) to plot
            title: Plot title
            figsize: Figure size
            **kwargs: Additional matplotlib arguments

        Returns:
            matplotlib Figure object

        Example:
            >>> fig = viz.plot_time_series(df, 'date', ['sales', 'profit'])
        """
        logger.info(f"[VISUALIZE] Creating time series plot")

        if isinstance(value_cols, str):
            value_cols = [value_cols]

        fig, ax = plt.subplots(figsize=figsize)

        for col in value_cols:
            ax.plot(df[date_col], df[col], label=col, **kwargs)

        ax.set_xlabel('Date')
        ax.set_ylabel('Value')
        ax.set_title(title or 'Time Series Plot')
        ax.legend()
        plt.xticks(rotation=45)
        plt.tight_layout()
        return fig


# ============================================================================
# Multi-Panel Figures
# ============================================================================

def create_dashboard(
    df: pd.DataFrame,
    title: str = "Data Dashboard",
    figsize: Tuple[int, int] = (16, 10)
) -> plt.Figure:
    """Create a multi-panel dashboard with multiple visualizations.

    Args:
        df: Input DataFrame
        title: Dashboard title
        figsize: Figure size

    Returns:
        matplotlib Figure object with subplots

    Example:
        >>> fig = create_dashboard(df, "Sales Dashboard")
        >>> fig.savefig('dashboard.png', dpi=300)
    """
    logger.info(f"[DASHBOARD] Creating dashboard: {title}")

    fig, axes = plt.subplots(2, 3, figsize=figsize)
    fig.suptitle(title, fontsize=16, fontweight='bold')

    viz = DataVisualizer()

    # Get numeric columns
    numeric_cols = df.select_dtypes(include=['number']).columns.tolist()

    # Plot 1: Histogram (first numeric column)
    if numeric_cols:
        ax = axes[0, 0]
        df[numeric_cols[0]].hist(bins=30, ax=ax)
        ax.set_title(f'Distribution: {numeric_cols[0]}')
        ax.set_xlabel(numeric_cols[0])

    # Plot 2: Box plots for numeric columns
    ax = axes[0, 1]
    if len(numeric_cols) > 0:
        df[numeric_cols[:5]].boxplot(ax=ax)
        ax.set_title('Box Plots')
        plt.xticks(rotation=45)

    # Plot 3: Correlation heatmap (if multiple numeric columns)
    ax = axes[0, 2]
    if len(numeric_cols) > 1:
        corr = df[numeric_cols].corr()
        sns.heatmap(corr, annot=True, fmt='.2f', cmap='coolwarm',
                   center=0, ax=ax, cbar=False)
        ax.set_title('Correlation')

    # Plot 4: Bar chart (first categorical column)
    ax = axes[1, 0]
    categorical_cols = df.select_dtypes(include=['object', 'category']).columns.tolist()
    if categorical_cols:
        df[categorical_cols[0]].value_counts().head(10).plot(kind='bar', ax=ax)
        ax.set_title(f'Count: {categorical_cols[0]}')
        plt.xticks(rotation=45)

    # Plot 5: Scatter plot (first two numeric columns)
    ax = axes[1, 1]
    if len(numeric_cols) >= 2:
        ax.scatter(df[numeric_cols[0]], df[numeric_cols[1]], alpha=0.5)
        ax.set_xlabel(numeric_cols[0])
        ax.set_ylabel(numeric_cols[1])
        ax.set_title(f'{numeric_cols[0]} vs {numeric_cols[1]}')

    # Plot 6: Statistics summary
    ax = axes[1, 2]
    ax.axis('off')
    summary_text = f"""
    Shape: {df.shape[0]} rows × {df.shape[1]} cols
    Memory: {df.memory_usage(deep=True).sum() / 1024:.1f} KB

    Numeric Columns: {len(numeric_cols)}
    Categorical Columns: {len(categorical_cols)}

    Missing Values: {df.isnull().sum().sum()}
    Duplicates: {df.duplicated().sum()}
    """
    ax.text(0.1, 0.5, summary_text, fontsize=10, family='monospace',
            verticalalignment='center')

    plt.tight_layout()
    return fig


# ============================================================================
# Report Integration
# ============================================================================

class ReportGeneratorWithCharts:
    """Enhanced report generator with charts integration.

    Example:
        >>> generator = ReportGeneratorWithCharts()
        >>> generator.generate_excel_with_charts(
        ...     df,
        ...     "report.xlsx",
        ...     charts=['histogram', 'boxplot', 'correlation']
        ... )
    """

    def __init__(self):
        """Initialize generator."""
        self.viz = DataVisualizer()

    def create_figure(
        self,
        df: pd.DataFrame,
        chart_type: str,
        **kwargs
    ) -> plt.Figure:
        """Create a specific type of chart.

        Args:
            df: Input DataFrame
            chart_type: Type of chart ('histogram', 'boxplot', 'scatter', etc.)
            **kwargs: Chart-specific arguments

        Returns:
            matplotlib Figure
        """
        chart_methods = {
            'histogram': self.viz.plot_histogram,
            'boxplot': self.viz.plot_boxplot,
            'violin': self.viz.plot_violin,
            'scatter': self.viz.plot_scatter,
            'bar': self.viz.plot_bar,
            'line': self.viz.plot_line,
            'correlation': self.viz.plot_correlation_heatmap,
            'timeseries': self.viz.plot_time_series,
        }

        if chart_type not in chart_methods:
            raise ValueError(f"Unsupported chart type: {chart_type}")

        return chart_methods[chart_type](df, **kwargs)

    def save_charts_to_pdf(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        charts: List[str],
        **kwargs
    ) -> None:
        """Generate multiple charts and save to PDF.

        Args:
            df: Input DataFrame
            output_path: Output PDF path
            charts: List of chart types to generate
            **kwargs: Additional arguments for charts
        """
        from matplotlib.backends.backend_pdf import PdfPages

        logger.info(f"[CHARTS] Creating PDF with charts: {charts}")

        with PdfPages(output_path) as pdf:
            for chart_type in charts:
                try:
                    fig = self.create_figure(df, chart_type, **kwargs)
                    pdf.savefig(fig)
                    plt.close(fig)
                    logger.info(f"[CHARTS] Added {chart_type} to PDF")
                except Exception as e:
                    logger.error(f"[CHARTS] Error creating {chart_type}: {e}")

        logger.info(f"[CHARTS] PDF saved to {output_path}")

    def generate_excel_with_charts(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        charts: Optional[List[str]] = None,
        **kwargs
    ) -> str:
        """Generate Excel report with embedded charts.

        Args:
            df: Input DataFrame
            output_path: Output Excel path
            charts: List of chart types to include
            **kwargs: Additional arguments

        Returns:
            Path to generated file
        """
        from core.reporter import generate_excel_report

        # Generate Excel
        logger.info(f"[CHARTS] Generating Excel with charts")
        generate_excel_report(df, output_path)

        # Create Excel with charts using XlsxWriter
        if charts:
            from openpyxl import load_workbook
            from openpyxl.drawing.image import Image
            from openpyxl.utils import get_column_letter

            # Generate charts as images
            chart_images = {}
            for chart_type in charts:
                try:
                    fig = self.create_figure(df, chart_type, **kwargs)

                    # Save figure to temporary image
                    img_path = Path(output_path).parent / f"chart_{chart_type}.png"
                    fig.savefig(img_path, dpi=150, bbox_inches='tight')
                    plt.close(fig)

                    chart_images[chart_type] = img_path
                except Exception as e:
                    logger.error(f"Error generating chart {chart_type}: {e}")

            # Insert images into Excel
            if chart_images:
                wb = load_workbook(output_path)
                ws = wb.create_sheet("Charts")

                row = 1
                for chart_type, img_path in chart_images.items():
                    # Add image
                    img = Image(img_path)
                    ws.add_image(img, f'{row}{get_column_letter(1)}')

                    # Add label
                    ws.cell(row + 20, 1, chart_type)

                    row += 25

                wb.save(output_path)
                logger.info("[CHARTS] Charts embedded in Excel")

        return str(output_path)
